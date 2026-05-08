#!/usr/bin/env python3
"""Import Retreat_Volunteer_Schedule_v5.xlsx into JewelHeart Postgres tables.

Reads multiple sheets from:
  docs/scheduling-reference/Retreat_Volunteer_Schedule_v5.xlsx

  - Sheet "1. List by Day-Slot": primary source for jobs, slots, tasks (same model as seed_summer_retreat_july_2026_from_xlsx.py).
  - Sheet "3. Site Matrix": optional cross-check (--verify-site-matrix) or extra tasks (--merge-site-matrix).
  - Sheet "4. Metrics": printed summary only (informational).
  - Sheet "5. Notes": optional export to text (--export-notes).

Requires: openpyxl, psql + DATABASE_URL for --apply.

Idempotent for a fixed retreat UUID (DELETE CASCADE then INSERT).

Usage:
  python3 scripts/import_retreat_volunteer_schedule_v5.py --dry-run > /tmp/jh-v5.sql
  psql "$DATABASE_URL" -v ON_ERROR_STOP=1 -f /tmp/jh-v5.sql

  python3 scripts/import_retreat_volunteer_schedule_v5.py --apply

  python3 scripts/import_retreat_volunteer_schedule_v5.py --apply --verify-site-matrix
  python3 scripts/import_retreat_volunteer_schedule_v5.py --export-notes
"""

from __future__ import annotations

import argparse
import os
import re
import subprocess
import sys
import uuid
from datetime import date
from pathlib import Path

NS = uuid.UUID("6ba7b810-9dad-11d1-80b4-00c04fd430c8")

# Slot name (first column of slot label before " — ") -> Postgres enum jewelheart_time_band
SLOT_TO_BAND: dict[str, str] = {
    "Start day": "early",
    "Morning break": "early",
    "Lunch break": "lunchtime",
    "Afternoon break": "anytime",
    "Dinner break": "dinnertime",
    "End day": "anytime",
    "All day": "allday",
}

MONTHS = {m: i for i, m in enumerate("Jan Feb Mar Apr May Jun Jul Aug Sep Oct Nov Dec".split(), 1)}

TASK_PARENS = re.compile(r"\s*\(\d+v,\s*\d+m\)\s*$", re.IGNORECASE)
# v5 day cell: "1: Mon Jul 20" or legacy "Day 1 ... Mon Jul 20"
DAY_V5 = re.compile(r"^\d+:\s*(Mon|Tue|Wed|Thu|Fri|Sat|Sun)\s+(\w{3})\s+(\d{1,2})\s*$")
DAY_LEGACY = re.compile(r"(Mon|Tue|Wed|Thu|Fri|Sat|Sun)\s+(\w{3})\s+(\d{1,2})")
MATRIX_LINE = re.compile(
    r"^(.+?):\s*(.+?)\s*\((\d+)v,\s*(\d+)m\)\s*$",
    re.IGNORECASE,
)
MATRIX_SHORT = re.compile(r"^(.+?)\s*\((\d+)v,\s*(\d+)m\)\s*$", re.IGNORECASE)
HEADER_DATE = re.compile(r"^\d+:\s*(Mon|Tue|Wed|Thu|Fri|Sat|Sun)\s+(\w{3})\s+(\d{1,2})")


def _u5(*parts: str) -> str:
    return str(uuid.uuid5(NS, ":".join(parts)))


def _esc(s: str) -> str:
    return s.replace("\\", "\\\\").replace("'", "''")


def _parse_day_to_iso(day_cell: str) -> str | None:
    s = str(day_cell).strip()
    m = DAY_V5.match(s)
    if not m and s.startswith("Day "):
        m = DAY_LEGACY.search(s)
    if not m:
        return None
    mon, d = m.group(2), int(m.group(3))
    if mon not in MONTHS:
        return None
    return f"2026-{MONTHS[mon]:02d}-{d:02d}"


def _parse_list_sheet(ws) -> list[tuple[str, str, str, str, str, int, int]]:
    """Rows: (slot_date, slot_name, time_range, site, full_task, volunteers, minutes)."""
    rows = list(ws.iter_rows(values_only=True))
    cur_date: str | None = None
    cur_slot: str | None = None
    cur_time: str | None = None
    out: list[tuple[str, str, str, str, str, int, int]] = []
    for row in rows[4:]:
        if not row or len(row) < 5:
            continue
        day_cell, slot_cell, time_cell, site, task = row[0], row[1], row[2], row[3], row[4]
        if day_cell:
            d = _parse_day_to_iso(str(day_cell))
            if d:
                cur_date = d
        if slot_cell:
            cur_slot = str(slot_cell).strip()
        if time_cell:
            cur_time = str(time_cell).strip()
        if not (site and task and cur_date and cur_slot):
            continue
        t = str(task).strip()
        m = re.search(r"\((\d+)v,\s*(\d+)m\)", t, re.IGNORECASE)
        if not m:
            continue
        vn, em = int(m.group(1)), int(m.group(2))
        out.append((cur_date, cur_slot, cur_time or "", str(site).strip(), t, vn, em))
    return out


def _parse_matrix_header(row) -> list[str | None]:
    """Row 4 of Site Matrix: Site / Job, then day headers -> ISO dates."""
    headers: list[str | None] = [None]
    for c in range(1, len(row)):
        cell = row[c]
        if cell is None:
            headers.append(None)
            continue
        s = str(cell).strip()
        m = HEADER_DATE.match(s)
        if not m:
            headers.append(None)
            continue
        mon, d = m.group(2), int(m.group(3))
        if mon not in MONTHS:
            headers.append(None)
            continue
        headers.append(f"2026-{MONTHS[mon]:02d}-{d:02d}")
    return headers


def _parse_matrix_line(line: str) -> tuple[str, str, int, int] | None:
    s = line.strip()
    if not s or s in ("—", "-"):
        return None
    m = MATRIX_LINE.match(s)
    if m:
        return m.group(1).strip(), m.group(2).strip(), int(m.group(3)), int(m.group(4))
    m2 = MATRIX_SHORT.match(s)
    if m2:
        return "All day", m2.group(1).strip(), int(m2.group(2)), int(m2.group(3))
    return None


def _parse_site_matrix(ws) -> list[tuple[str, str, str, str, int, int]]:
    """Rows: (slot_date, slot_name, time_range, site, full_task, vn, em) compatible with build_sql."""
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 5:
        return []
    header_row = rows[3]
    dates = _parse_matrix_header(header_row)
    out: list[tuple[str, str, str, str, str, int, int]] = []
    for row in rows[4:]:
        if not row or not row[0]:
            continue
        site = str(row[0]).strip()
        for col_idx in range(1, len(dates)):
            if col_idx >= len(row):
                break
            slot_date = dates[col_idx]
            if not slot_date:
                continue
            cell = row[col_idx]
            if cell is None:
                continue
            text = str(cell).strip()
            if not text:
                continue
            for line in text.splitlines():
                parsed = _parse_matrix_line(line)
                if not parsed:
                    continue
                slot_prefix, desc, vn, em = parsed
                full_task = f"{desc} ({vn}v, {em}m)"
                out.append((slot_date, slot_prefix, "", site, full_task, vn, em))
    return out


def _job_title(site: str, full_task: str) -> str:
    base = TASK_PARENS.sub("", full_task).strip()
    return f"{site} — {base}"


def build_sql(
    *,
    retreat_id: str,
    parsed: list[tuple[str, str, str, str, str, int, int]],
    retreat_name: str,
    tz: str,
    start_date: str,
    end_date: str,
) -> str:
    if not parsed:
        raise SystemExit("No schedule rows parsed (list sheet empty?).")

    job_specs: dict[tuple[str, int, int], str] = {}
    for _, _, _, site, full_task, vn, em in parsed:
        title = _job_title(site, full_task)
        key = (title, vn, em)
        if key not in job_specs:
            job_specs[key] = _u5("JH:job", retreat_id, title, str(vn), str(em))

    slot_specs: dict[tuple[str, str, str], tuple[str, str, str, str]] = {}
    for slot_date, slot_name, time_range, _, _, _, _ in parsed:
        sk = (slot_date, slot_name, time_range)
        if sk in slot_specs:
            continue
        band = SLOT_TO_BAND.get(slot_name)
        if not band:
            raise SystemExit(f"Unknown slot name (add SLOT_TO_BAND mapping): {slot_name!r}")
        sid = _u5("JH:slot", retreat_id, slot_date, slot_name, time_range)
        label = f"{slot_name} — {time_range}" if time_range else slot_name
        d = date.fromisoformat(slot_date)
        dow = d.strftime("%A")
        slot_specs[sk] = (sid, label, band, dow)

    lines: list[str] = [
        "BEGIN;",
        f"DELETE FROM jewelheart_retreats WHERE id = '{retreat_id}'::uuid;",
        "INSERT INTO jewelheart_retreats (id, name, timezone, start_date, end_date, status, created_at, updated_at)",
        "VALUES (",
        f"  '{retreat_id}'::uuid,",
        f"  '{_esc(retreat_name)}',",
        f"  '{_esc(tz)}',",
        f"  '{start_date}'::date,",
        f"  '{end_date}'::date,",
        "  'published',",
        "  now(),",
        "  now()",
        ");",
        "",
    ]

    uid = (os.environ.get("JEWELHEART_SEED_FIREBASE_UID") or "").strip()
    if uid:
        lines.extend(
            [
                "INSERT INTO jewelheart_retreat_admins (retreat_id, firebase_uid, created_at)",
                f"VALUES ('{retreat_id}'::uuid, '{_esc(uid)}', now())",
                "ON CONFLICT (retreat_id, firebase_uid) DO NOTHING;",
                "",
            ]
        )

    lines.append("-- jobs")
    for (title, vn, em), jid in sorted(job_specs.items(), key=lambda x: x[0][0]):
        lines.append(
            "INSERT INTO jewelheart_jobs (id, retreat_id, title, volunteers_needed, estimated_minutes, created_at, updated_at) "
            f"VALUES ('{jid}'::uuid, '{retreat_id}'::uuid, '{_esc(title)}', {vn}, {em}, now(), now());"
        )
    lines.append("")
    lines.append("-- slots")
    for sk in sorted(slot_specs.keys(), key=lambda x: (x[0], x[1], x[2])):
        sid, label, band, dow = slot_specs[sk]
        slot_date, _, _ = sk
        lines.append(
            "INSERT INTO jewelheart_slots (id, retreat_id, label, slot_date, day_of_week, activity_context, time_band, created_at, updated_at) "
            f"VALUES ('{sid}'::uuid, '{retreat_id}'::uuid, '{_esc(label)}', '{slot_date}'::date, '{dow}', NULL, '{band}'::jewelheart_time_band, now(), now());"
        )
    lines.append("")
    lines.append("-- tasks (job × slot)")
    seen_task: set[tuple[str, str]] = set()
    for slot_date, slot_name, time_range, site, full_task, vn, em in parsed:
        title = _job_title(site, full_task)
        jid = job_specs[(title, vn, em)]
        sk = (slot_date, slot_name, time_range)
        sid = slot_specs[sk][0]
        pair = (jid, sid)
        if pair in seen_task:
            continue
        seen_task.add(pair)
        tid = _u5("JH:task", retreat_id, jid, sid)
        lines.append(
            "INSERT INTO jewelheart_tasks (id, retreat_id, job_id, slot_id, notes, created_at, updated_at) "
            f"VALUES ('{tid}'::uuid, '{retreat_id}'::uuid, '{jid}'::uuid, '{sid}'::uuid, NULL, now(), now());"
        )
    lines.append("COMMIT;")
    return "\n".join(lines) + "\n"


def _slot_time_lookup(
    list_rows: list[tuple[str, str, str, str, str, int, int]],
) -> dict[tuple[str, str], str]:
    """Map (slot_date, slot_name) -> time_range from list sheet (for merging matrix rows)."""
    m: dict[tuple[str, str], str] = {}
    for slot_date, slot_name, time_range, *_ in list_rows:
        m[(slot_date, slot_name)] = time_range
    return m


def _enrich_matrix_rows(
    list_rows: list[tuple[str, str, str, str, str, int, int]],
    matrix_rows: list[tuple[str, str, str, str, str, int, int]],
) -> list[tuple[str, str, str, str, str, int, int]]:
    """Fill empty matrix time_range from list sheet so slots share one row per (date, slot name)."""
    lu = _slot_time_lookup(list_rows)
    out: list[tuple[str, str, str, str, str, int, int]] = []
    for slot_date, slot_name, time_range, site, full_task, vn, em in matrix_rows:
        tr = time_range or lu.get((slot_date, slot_name), "")
        out.append((slot_date, slot_name, tr, site, full_task, vn, em))
    return out


def _merge_parsed(
    a: list[tuple[str, str, str, str, str, int, int]],
    b: list[tuple[str, str, str, str, str, int, int]],
) -> list[tuple[str, str, str, str, str, int, int]]:
    """Union by logical task key (dedupe same job+slot date+name+time+vn+em)."""
    seen: set[tuple[str, str, str, str, int, int]] = set()
    out: list[tuple[str, str, str, str, str, int, int]] = []
    for row in a + b:
        slot_date, slot_name, time_range, site, full_task, vn, em = row
        title = _job_title(site, full_task)
        key = (slot_date, slot_name, time_range, title, vn, em)
        if key in seen:
            continue
        seen.add(key)
        out.append(row)
    return out


def _verify_site_matrix(
    list_rows: list[tuple[str, str, str, str, str, int, int]],
    matrix_rows: list[tuple[str, str, str, str, str, int, int]],
) -> list[str]:
    """Return warning strings (not fatal)."""
    warnings: list[str] = []
    # Compare on date + slot name + job title + v + m (ignore time string differences).
    def key(r: tuple[str, str, str, str, str, int, int]) -> tuple[str, str, str, int, int]:
        return (r[0], r[1], _job_title(r[3], r[4]), r[5], r[6])

    set_list = {key(r) for r in list_rows}
    mat_enriched = _enrich_matrix_rows(list_rows, matrix_rows)
    set_mat = {key(r) for r in mat_enriched}
    only_mat = set_mat - set_list
    only_list = set_list - set_mat
    if only_mat:
        warnings.append(f"Site matrix has {len(only_mat)} logical row(s) not found in list sheet (sample): {next(iter(only_mat))}")
    if only_list:
        warnings.append(f"List sheet has {len(only_list)} logical row(s) not found in site matrix (sample): {next(iter(only_list))}")
    return warnings


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument(
        "--xlsx",
        type=Path,
        default=Path(__file__).resolve().parents[1]
        / "docs/scheduling-reference/Retreat_Volunteer_Schedule_v5.xlsx",
    )
    ap.add_argument("--dry-run", action="store_true", help="Print SQL to stdout (default if no --apply).")
    ap.add_argument("--apply", action="store_true", help="Run SQL via psql and DATABASE_URL.")
    ap.add_argument(
        "--merge-site-matrix",
        action="store_true",
        help="Include tasks parsed from sheet 3 (Site Matrix); merges with list sheet with dedupe.",
    )
    ap.add_argument(
        "--verify-site-matrix",
        action="store_true",
        help="Compare sheet 1 vs sheet 3 and print warnings to stderr (no SQL change).",
    )
    ap.add_argument(
        "--export-notes",
        action="store_true",
        help="Write sheet 5 (Notes) to docs/scheduling-reference/Retreat_Volunteer_Schedule_v5_notes_export.txt",
    )
    ap.add_argument("--retreat-id-seed", default="summer-2026-07-20-v5", help="Seed string for stable UUID namespace.")
    args = ap.parse_args()

    try:
        import openpyxl
    except ImportError:
        print("Install openpyxl: pip install openpyxl", file=sys.stderr)
        raise SystemExit(1)

    if not args.xlsx.is_file():
        raise SystemExit(f"Missing workbook: {args.xlsx}")

    wb = openpyxl.load_workbook(args.xlsx, read_only=True, data_only=True)
    try:
        ws1 = wb["1. List by Day-Slot"]
        list_parsed = _parse_list_sheet(ws1)
        retreat_title = str(ws1.cell(row=1, column=1).value or "").strip() or "Retreat July 20-26, 2026 (v5)"

        matrix_parsed: list[tuple[str, str, str, str, str, int, int]] = []
        if "3. Site Matrix" in wb.sheetnames:
            matrix_parsed = _parse_site_matrix(wb["3. Site Matrix"])

        metrics_lines: list[str] = []
        if "4. Metrics" in wb.sheetnames:
            mws = wb["4. Metrics"]
            for i, row in enumerate(mws.iter_rows(min_row=1, max_row=12, values_only=True)):
                metrics_lines.append("  " + " | ".join("" if c is None else str(c) for c in row))

        notes_text = ""
        if "5. Notes" in wb.sheetnames:
            nws = wb["5. Notes"]
            chunks = []
            for row in nws.iter_rows(values_only=True):
                line = " | ".join("" if c is None else str(c).strip() for c in row if c is not None)
                if line.strip():
                    chunks.append(line)
            notes_text = "\n".join(chunks)
    finally:
        wb.close()

    if args.export_notes:
        out_notes = (
            Path(__file__).resolve().parents[1]
            / "docs/scheduling-reference/Retreat_Volunteer_Schedule_v5_notes_export.txt"
        )
        out_notes.write_text(notes_text or "(empty notes sheet)\n", encoding="utf-8")
        sys.stderr.write(f"Wrote {out_notes}\n")

    notes_only = (
        args.export_notes
        and not args.apply
        and not args.merge_site_matrix
        and not args.verify_site_matrix
    )
    if notes_only:
        return

    if args.verify_site_matrix and matrix_parsed:
        for w in _verify_site_matrix(list_parsed, matrix_parsed):
            sys.stderr.write(f"VERIFY: {w}\n")

    if args.merge_site_matrix and matrix_parsed:
        matrix_enriched = _enrich_matrix_rows(list_parsed, matrix_parsed)
        parsed = _merge_parsed(list_parsed, matrix_enriched)
    else:
        parsed = list_parsed

    if not parsed:
        raise SystemExit("No rows after parsing. Check sheet '1. List by Day-Slot'.")

    retreat_id = _u5("JH:retreat", args.retreat_id_seed)
    dates = sorted({r[0] for r in parsed})
    start_date = dates[0]
    # Workbook title spans Jul 20–26; day 7 has no volunteer tasks but end_date covers the week.
    end_date = max(dates + ["2026-07-26"])

    sql = build_sql(
        retreat_id=retreat_id,
        parsed=parsed,
        retreat_name=retreat_title[:200],
        tz="America/Chicago",
        start_date=start_date,
        end_date=end_date,
    )

    if metrics_lines:
        sys.stderr.write("-- Metrics sheet (informational)\n")
        sys.stderr.write("\n".join(metrics_lines) + "\n\n")

    if args.apply:
        db = (os.environ.get("DATABASE_URL") or "").strip()
        if not db:
            raise SystemExit("DATABASE_URL must be set for --apply")
        r = subprocess.run(
            ["psql", db, "-v", "ON_ERROR_STOP=1", "-f", "-"],
            input=sql.encode(),
            capture_output=True,
        )
        if r.returncode != 0:
            sys.stderr.buffer.write(r.stderr or b"")
            sys.stdout.buffer.write(r.stdout or b"")
            raise SystemExit(r.returncode)
        sys.stderr.write(
            f"OK: imported retreat id {retreat_id} ({len(parsed)} task rows after merge/dedupe).\n"
        )
        return

    sys.stdout.write(sql)


if __name__ == "__main__":
    main()
